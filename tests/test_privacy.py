"""
Приватность данных: скрытые цели, шифрование полей, мягкое удаление,
журнал доступа (ТЗ 5.2, 8.1, 8.4, 9.5).
"""
from __future__ import annotations

import datetime as dt
from io import BytesIO

import pytest
from django.urls import reverse
from openpyxl import load_workbook

from apps.core.audit import AuditAction
from apps.core.models import AuditLog
from apps.core.tenancy import organization_context
from apps.journal.models import Goal, GoalKind, GoalVisibility, MoodEntry, Student
from apps.journal.services.exports import goals_xlsx
from tests.conftest import login


@pytest.fixture
def goals(tenant_a):
    with organization_context(tenant_a.organization):
        open_goal = Goal.objects.create(
            organization=tenant_a.organization, student=tenant_a.student,
            kind=GoalKind.PERSONAL, visibility=GoalVisibility.OPEN,
            title="Разобраться с тригонометрией",
        )
        hidden_goal = Goal.objects.create(
            organization=tenant_a.organization, student=tenant_a.student,
            kind=GoalKind.PERSONAL, visibility=GoalVisibility.HIDDEN,
            title="Секретная цель про поступление",
        )
    return open_goal, hidden_goal


def test_hidden_goals_never_reach_exports(tenant_a, goals):
    """Ключевое требование: скрытая цель не попадает ни в одну выгрузку."""
    open_goal, hidden_goal = goals
    with organization_context(tenant_a.organization):
        payload = goals_xlsx(tenant_a.student)
    sheet = load_workbook(BytesIO(payload)).active
    titles = {row[2] for row in sheet.iter_rows(min_row=2, values_only=True)}
    assert open_goal.title in titles
    assert hidden_goal.title not in titles


def test_hidden_goals_not_visible_to_parent(client, tenant_a, goals):
    open_goal, hidden_goal = goals
    login(client, tenant_a, tenant_a.parent_user)
    body = client.get(reverse("cabinet:parent_home")).content.decode()
    assert hidden_goal.title not in body


def test_hidden_goals_visible_to_student_only(client, tenant_a, goals):
    open_goal, hidden_goal = goals
    login(client, tenant_a, tenant_a.student_user)
    body = client.get(reverse("cabinet:student_home")).content.decode()
    assert hidden_goal.title in body
    assert open_goal.title in body


def test_visible_to_others_manager_excludes_hidden(tenant_a, goals):
    with organization_context(tenant_a.organization):
        visible = list(Goal.objects.visible_to_others().values_list("visibility", flat=True))
    assert visible == [GoalVisibility.OPEN]


def test_mood_aggregate_is_not_public(client_a, tenant_a):
    """Настроение группы — внутренний инструмент, наружу не выводится."""
    with organization_context(tenant_a.organization):
        MoodEntry.objects.create(
            organization=tenant_a.organization, student=tenant_a.student,
            day=dt.date.today(), value=MoodEntry.Scale.LOW,
        )
    body = client_a.get(reverse("public:landing")).content.decode().lower()
    assert "настроение класса" not in body
    assert "индикатор состояния" not in body


def test_birth_date_is_encrypted_at_rest(tenant_a):
    """В базе лежит шифротекст, в Python — дата."""
    from django.db import connection

    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT birth_date FROM journal_student WHERE id = %s", [str(tenant_a.student.pk)]
        )
        raw = cursor.fetchone()[0]
    assert raw.startswith("enc$")
    assert "2010" not in raw

    with organization_context(tenant_a.organization):
        student = Student.objects.get(pk=tenant_a.student.pk)
    assert student.birth_date == dt.date(2010, 3, 14)


def test_student_soft_delete_and_restore(client, tenant_a):
    """Случайно удалённого ученика можно вернуть посреди учебного года."""
    with organization_context(tenant_a.organization):
        tenant_a.student.delete()
        assert Student.objects.filter(pk=tenant_a.student.pk).count() == 0
        assert Student.all_objects.filter(pk=tenant_a.student.pk).count() == 1

    login(client, tenant_a, tenant_a.owner_user)
    response = client.post(reverse("cabinet:student_restore", args=[tenant_a.student.pk]))
    assert response.status_code == 200

    with organization_context(tenant_a.organization):
        assert Student.objects.filter(pk=tenant_a.student.pk).count() == 1


def test_viewing_student_card_is_audited(client, tenant_a):
    login(client, tenant_a, tenant_a.parent_user)
    client.get(reverse("cabinet:parent_home"))
    assert AuditLog.objects.filter(
        action=AuditAction.VIEW_STUDENT, object_id=str(tenant_a.student.pk)
    ).exists()


def test_export_is_audited(client, tenant_a):
    login(client, tenant_a, tenant_a.owner_user)
    client.get(reverse("cabinet:export_students"))
    entry = AuditLog.objects.filter(action=AuditAction.EXPORT).first()
    assert entry is not None
    assert entry.extra["export"] == "students"


def test_login_is_audited(client, tenant_a):
    client.defaults["HTTP_HOST"] = tenant_a.host
    client.post(
        reverse("accounts:login"),
        {"username": tenant_a.parent_user.email, "password": "test-parol-12345"},
    )
    assert AuditLog.objects.filter(action=AuditAction.LOGIN).exists()


def test_failed_login_is_audited(client, tenant_a):
    client.defaults["HTTP_HOST"] = tenant_a.host
    client.post(
        reverse("accounts:login"),
        {"username": tenant_a.parent_user.email, "password": "wrong"},
    )
    assert AuditLog.objects.filter(action=AuditAction.LOGIN_FAILED).exists()


def test_students_export_omits_sensitive_fields(tenant_a):
    from apps.journal.services.exports import students_xlsx

    with organization_context(tenant_a.organization):
        payload = students_xlsx()
    sheet = load_workbook(BytesIO(payload)).active
    headers = [cell.value for cell in sheet[1]]
    assert "Дата рождения" not in headers
    assert "Документ" not in headers
