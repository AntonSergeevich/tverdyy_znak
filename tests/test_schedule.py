"""Расписание в кабинете и загрузка недельной сетки (ТЗ 5.2, 5.3)."""
from __future__ import annotations

import datetime as dt

import pytest
from django.core.management import call_command
from django.core.management.base import CommandError
from django.urls import reverse
from django.utils import timezone

from apps.journal.models import Lesson
from tests.conftest import PASSWORD


def _login(client, tenant, user):
    client.defaults["HTTP_HOST"] = tenant.host
    client.post(reverse("accounts:login"), {"username": user.email, "password": PASSWORD})


def test_schedule_page_shows_own_lessons(client, tenant_a):
    _login(client, tenant_a, tenant_a.student_user)
    response = client.get(reverse("cabinet:schedule"))

    assert response.status_code == 200
    body = response.content.decode()
    assert tenant_a.subject.name in body
    assert "Расписание" in body


def test_schedule_page_hides_other_tenant(client, tenant_a, tenant_b):
    """Ученик одной организации не видит занятий другой — даже на общей странице."""
    _login(client, tenant_a, tenant_a.student_user)
    body = client.get(reverse("cabinet:schedule")).content.decode()

    assert tenant_b.group.name not in body


def test_schedule_page_shows_empty_week_without_crashing(client, tenant_a):
    _login(client, tenant_a, tenant_a.student_user)
    response = client.get(reverse("cabinet:schedule"), {"n": "40"})

    assert response.status_code == 200
    assert "занятий пока нет" in response.content.decode()


def test_schedule_week_offset_is_bounded(client, tenant_a):
    """Мусор в параметре недели не должен ронять страницу."""
    _login(client, tenant_a, tenant_a.student_user)

    assert client.get(reverse("cabinet:schedule"), {"n": "нет"}).status_code == 200
    assert client.get(reverse("cabinet:schedule"), {"n": "99999"}).status_code == 200


def test_schedule_link_to_source_file_appears_when_set(client, tenant_a):
    tenant_a.organization.schedule_url = "https://docs.yandex.ru/view/d/example"
    tenant_a.organization.save(update_fields=["schedule_url"])

    _login(client, tenant_a, tenant_a.parent_user)
    body = client.get(reverse("cabinet:schedule"), {"n": "40"}).content.decode()

    assert "https://docs.yandex.ru/view/d/example" in body


def _write_grid(tmp_path, tenant, rows):
    path = tmp_path / "grid.csv"
    lines = ["day,time,subject,group,teacher,room,duration"] + rows
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


@pytest.fixture
def grid(tmp_path, tenant_a):
    return _write_grid(
        tmp_path, tenant_a,
        [f"пн,09:00,{tenant_a.subject.name},{tenant_a.group.name},,каб. 2,45"],
    )


def test_import_schedule_expands_week_over_module(grid, tenant_a):
    """Одна строка сетки — занятие на каждый понедельник модуля."""
    call_command("import_schedule", str(grid), "--organization", tenant_a.organization.slug,
                 "--module", tenant_a.module.number)

    lessons = Lesson.all_objects.filter(
        organization=tenant_a.organization, module=tenant_a.module, room="каб. 2"
    )
    assert lessons.count() > 1
    for lesson in lessons:
        local = lesson.starts_at.astimezone(tenant_a.organization.tzinfo)
        assert local.weekday() == 0
        assert (local.hour, local.minute) == (9, 0)
        assert tenant_a.module.starts_on <= local.date() <= tenant_a.module.ends_on


def test_import_schedule_keeps_local_time_of_the_centre(grid, tenant_a):
    """
    09:00 в сетке — это 09:00 в Красноярске, а не на сервере.

    Сервер живёт в UTC, и без явного часового пояса организации занятия
    уезжали бы на семь часов вперёд.
    """
    call_command("import_schedule", str(grid), "--organization", tenant_a.organization.slug,
                 "--module", tenant_a.module.number)

    lesson = Lesson.all_objects.filter(
        organization=tenant_a.organization, module=tenant_a.module, room="каб. 2"
    ).first()
    local = lesson.starts_at.astimezone(tenant_a.organization.tzinfo)
    assert (local.hour, local.minute) == (9, 0)


def test_import_schedule_is_idempotent(grid, tenant_a):
    args = (str(grid),)
    kwargs = {"organization": tenant_a.organization.slug, "module": tenant_a.module.number}
    call_command("import_schedule", *args, **kwargs)
    first = Lesson.all_objects.filter(organization=tenant_a.organization).count()
    call_command("import_schedule", *args, **kwargs)

    assert Lesson.all_objects.filter(organization=tenant_a.organization).count() == first


def test_import_schedule_dry_run_writes_nothing(grid, tenant_a):
    before = Lesson.all_objects.filter(organization=tenant_a.organization).count()
    call_command("import_schedule", str(grid), "--organization", tenant_a.organization.slug,
                 "--module", tenant_a.module.number, "--dry-run")

    assert Lesson.all_objects.filter(organization=tenant_a.organization).count() == before


def test_import_schedule_reports_missing_columns(tmp_path, tenant_a):
    path = tmp_path / "bad.csv"
    path.write_text("день,время\nпн,09:00", encoding="utf-8")

    with pytest.raises(CommandError, match="нет колонок"):
        call_command("import_schedule", str(path), "--organization", tenant_a.organization.slug)


def test_import_schedule_rejects_unknown_weekday(tmp_path, tenant_a):
    path = _write_grid(tmp_path, tenant_a,
                       [f"вторая,09:00,{tenant_a.subject.name},{tenant_a.group.name},,,45"])

    with pytest.raises(CommandError, match="день недели"):
        call_command("import_schedule", str(path), "--organization", tenant_a.organization.slug,
                     "--module", tenant_a.module.number)


def test_import_schedule_skips_unknown_subject_without_creating_it(tmp_path, tenant_a, capsys):
    """
    Опечатка в названии предмета не заводит новый предмет.

    Иначе в журнале рядом с «Алгеброй» появилась бы «алгебра ».
    """
    path = _write_grid(tmp_path, tenant_a,
                       [f"пн,09:00,Астрология,{tenant_a.group.name},,,45"])
    call_command("import_schedule", str(path), "--organization", tenant_a.organization.slug,
                 "--module", tenant_a.module.number)

    from apps.journal.models import Subject

    assert not Subject.all_objects.filter(name__iexact="Астрология").exists()
    assert "не найден" in capsys.readouterr().err


def test_example_grid_matches_expected_columns():
    """Образец в docs не должен разъехаться с тем, что читает команда."""
    from pathlib import Path

    from apps.journal.management.commands.import_schedule import (
        OPTIONAL_COLUMNS,
        REQUIRED_COLUMNS,
    )

    header = Path("docs/schedule.example.csv").read_text(encoding="utf-8").splitlines()[0]
    assert header.split(",") == REQUIRED_COLUMNS + OPTIONAL_COLUMNS


# ── разбор таблицы, в которой расписание ведёт заказчик ─────────────────────

def test_time_range_with_leading_hour_is_not_mistaken_for_lesson_number():
    """
    «9.30-10.10» начинается так же, как «9. …» — номер урока.

    Если снимать номер вслепую, от времени остаётся «30-10.10»,
    и весь лист молча читается как пустой.
    """
    from apps.journal.services.schedule_import import parse_time_range

    assert parse_time_range("9.30-10.10") == (dt.time(9, 30), 40)
    assert parse_time_range("1. 9.30-10.10") == (dt.time(9, 30), 40)
    assert parse_time_range("13:30-14:00") == (dt.time(13, 30), 30)
    assert parse_time_range("Время") is None


@pytest.mark.parametrize(
    "value, expected",
    [
        ("ПН 30.08", (0, 30, 8)),
        ("ВТ 01.09", (1, 1, 9)),
        ("02.09 Ср", (2, 2, 9)),
        ("Время", None),
        (None, None),
    ],
)
def test_day_header_parsing(value, expected):
    from apps.journal.services.schedule_import import parse_day_header

    assert parse_day_header(value) == expected


def test_alias_is_only_a_fallback():
    """
    Синоним не должен подменять предмет, который в журнале уже есть.

    «ВиСТ» разворачивается, но искать надо сначала по тому, что написано:
    у другой организации может быть предмет ровно с таким названием.
    """
    from apps.journal.services.schedule_import import title_candidates

    assert title_candidates("ВиСТ") == ["ВиСТ", "Вероятность и статистика"]
    assert title_candidates("  Химия  ") == ["Химия"]


def test_grid_parsing_reports_wrong_weekday_and_copied_header():
    """
    Две ошибки, которые реально были в файле заказчика.

    Дата с чужим днём недели и скопированная шапка второй недели: обе
    молча теряют занятия, поэтому обе должны быть названы вслух.
    """
    from apps.journal.services.schedule_import import parse_grid

    rows = [
        ["Время", "ПН 30.08", "ВТ 01.09"],
        ["9.30-10.10", "Русский язык", "Алгебра"],
        [None, None, None],
        ["Время", "ПН 07.09", "ВТ 01.09"],
        ["9.30-10.10", "Литература", "Геометрия"],
    ]
    result = parse_grid(rows, within=(dt.date(2026, 9, 1), dt.date(2027, 5, 21)))

    assert any("воскресенье" in w for w in result.warnings)
    assert any("скопировали" in w for w in result.warnings)
    assert {lesson.title for lesson in result.lessons} == {
        "Русский язык", "Алгебра", "Литература", "Геометрия"
    }


def test_workbook_from_yandex_docs_opens_despite_invalid_style(tmp_path):
    """
    Яндекс.Документы пишут в границах style="solid" — такого значения нет.

    openpyxl из-за этого отказывается открывать файл целиком, поэтому копия
    чинится на лету. Иначе заказчику пришлось бы пересохранять файл вручную.
    """
    import io
    import zipfile

    import openpyxl

    from apps.journal.services.schedule_import import load_workbook_rows

    book = openpyxl.Workbook()
    book.active.title = "Расписание"
    book.active["A1"] = "Время"
    buffer = io.BytesIO()
    book.save(buffer)

    broken = tmp_path / "yandex.xlsx"
    with zipfile.ZipFile(io.BytesIO(buffer.getvalue())) as zin:
        with zipfile.ZipFile(broken, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/styles.xml":
                    data = data.replace(
                        b"<borders", b'<borders><border><left style="solid"/></border>', 1
                    ).replace(b'count="1"', b'count="2"', 1)
                zout.writestr(item, data)

    with pytest.raises(ValueError):
        openpyxl.load_workbook(broken)
    assert load_workbook_rows(broken)["Расписание"][0][0] == "Время"


def _make_table(tmp_path, tenant, *, extra_rows=None):
    import openpyxl

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Расписание СЕНТЯБРЬ"
    monday = tenant.module.starts_on
    monday -= dt.timedelta(days=monday.weekday())
    if monday < tenant.module.starts_on:
        monday += dt.timedelta(days=7)
    sheet.append(["Время", f"ПН {monday:%d.%m}", f"ВТ {monday + dt.timedelta(days=1):%d.%m}"])
    sheet.append(["9.30-10.10", tenant.subject.name, tenant.subject.name])
    for row in extra_rows or []:
        sheet.append(row)
    path = tmp_path / "raspisanie.xlsx"
    book.save(path)
    return path, monday


def test_import_from_table_creates_lessons_on_listed_days(tmp_path, tenant_a):
    path, monday = _make_table(tmp_path, tenant_a)
    call_command("import_schedule", str(path), "--organization", tenant_a.organization.slug,
                 "--module", tenant_a.module.number, "--group", tenant_a.group.name)

    days = {
        lesson.starts_at.astimezone(tenant_a.organization.tzinfo).date()
        for lesson in Lesson.all_objects.filter(
            organization=tenant_a.organization, subject=tenant_a.subject
        )
    }
    assert monday in days
    assert monday + dt.timedelta(days=1) in days


def test_import_from_table_repeats_last_week_only_when_asked(tmp_path, tenant_a):
    """
    Продлевать расписание до конца модуля — осознанное решение.

    По умолчанию загружается ровно то, что в файле: додуманные занятия
    в кабинете ребёнка хуже, чем их отсутствие.
    """
    path, _ = _make_table(tmp_path, tenant_a)
    kwargs = {
        "organization": tenant_a.organization.slug,
        "module": tenant_a.module.number,
        "group": tenant_a.group.name,
    }
    call_command("import_schedule", str(path), **kwargs)
    without = Lesson.all_objects.filter(organization=tenant_a.organization).count()

    call_command("import_schedule", str(path), repeat_last_week=True, **kwargs)
    with_repeat = Lesson.all_objects.filter(organization=tenant_a.organization).count()

    assert with_repeat > without


def test_import_from_table_names_what_it_could_not_match(tmp_path, tenant_a, capsys):
    path, _ = _make_table(
        tmp_path, tenant_a, extra_rows=[["10.20-11.00", "Астрология", "Астрология"]]
    )
    call_command("import_schedule", str(path), "--organization", tenant_a.organization.slug,
                 "--module", tenant_a.module.number, "--group", tenant_a.group.name)

    err = capsys.readouterr().err
    assert "Астрология" in err


def test_day_blocks_are_not_graded(tenant_a):
    """
    Обед и утренний круг стоят в расписании, но 100 баллов по ним не раскладываются.

    Иначе педагог однажды заведёт структуру оценивания для обеда,
    и лимит модуля начнёт съедаться неучебным блоком.
    """
    from django.core.exceptions import ValidationError

    from apps.journal.models import Subject, SubjectKind
    from apps.journal.services.grading import create_default_structure

    lunch = Subject.all_objects.create(
        organization=tenant_a.organization, academic_year=tenant_a.year,
        name="Обед", kind=SubjectKind.ACTIVITY, weekly_hours=0,
    )
    assert lunch.is_graded is False
    with pytest.raises(ValidationError):
        create_default_structure(tenant_a.module, lunch, tenant_a.group)
