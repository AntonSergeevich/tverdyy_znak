"""
Изоляция организаций (ТЗ 3.1, 9.4).

Главный тест проекта: пользователь организации А не должен увидеть
ни одной записи организации Б ни через один эндпоинт.
"""
from __future__ import annotations

import pytest
from django.urls import reverse

from apps.core.tenancy import get_current_organization, organization_context, unscoped
from apps.journal.models import Lesson, Student
from tests.conftest import login


def test_manager_filters_by_current_organization(tenant_a, tenant_b):
    with organization_context(tenant_a.organization):
        assert list(Student.objects.values_list("pk", flat=True)) == [tenant_a.student.pk]
    with organization_context(tenant_b.organization):
        assert list(Student.objects.values_list("pk", flat=True)) == [tenant_b.student.pk]


def test_manager_returns_nothing_without_organization(tenant_a, tenant_b):
    """Без организации в контексте выборка пуста — лучше пусто, чем чужое."""
    assert get_current_organization() is None
    assert Student.objects.count() == 0
    assert Lesson.objects.count() == 0


def test_all_objects_sees_everything_only_explicitly(tenant_a, tenant_b):
    assert Student.all_objects.count() == 2
    with unscoped():
        assert Student.objects.count() == 2


UNSAFE_ENDPOINTS = [
    ("cabinet:parent_child", "student"),
    ("cabinet:lesson_journal", "lesson"),
]


@pytest.mark.parametrize("url_name, attribute", UNSAFE_ENDPOINTS)
def test_cross_tenant_object_access_is_denied(client, tenant_a, tenant_b, url_name, attribute):
    """Подстановка чужого id в URL не даёт доступа."""
    login(client, tenant_a, tenant_a.owner_user)
    foreign = getattr(tenant_b, attribute)
    kwargs = {"student_id" if attribute == "student" else "lesson_id": foreign.pk}
    response = client.get(reverse(url_name, kwargs=kwargs))
    assert response.status_code in (403, 404)


def test_lists_do_not_leak_between_tenants(client, tenant_a, tenant_b):
    login(client, tenant_a, tenant_a.owner_user)
    for url in [
        reverse("cabinet:students"),
        reverse("cabinet:leads"),
        reverse("cabinet:dashboard"),
        reverse("cabinet:payroll"),
    ]:
        body = client.get(url).content.decode()
        assert tenant_b.student.last_name not in body
        assert tenant_b.organization.name not in body


def test_export_contains_only_own_students(client, tenant_a, tenant_b):
    from apps.journal.services.exports import students_xlsx
    from openpyxl import load_workbook
    from io import BytesIO

    with organization_context(tenant_a.organization):
        payload = students_xlsx()
    sheet = load_workbook(BytesIO(payload)).active
    names = {row[1] for row in sheet.iter_rows(min_row=2, values_only=True)}
    assert tenant_a.student.first_name in names
    assert tenant_b.student.first_name not in names


def test_unknown_host_gets_no_organization(client, tenant_a):
    client.defaults["HTTP_HOST"] = "unknown.example.com"
    response = client.get("/")
    assert response.status_code == 404


def test_organization_resolved_by_domain(client, tenant_a, tenant_b):
    client.defaults["HTTP_HOST"] = tenant_b.host
    response = client.get("/")
    assert response.status_code == 200
    assert response.wsgi_request.organization == tenant_b.organization
