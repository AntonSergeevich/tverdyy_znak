"""
Разбор календарно-тематического планирования (КТП).

КТП приходит файлом — таблицей, которую педагог составлял не здесь.
Единого формата у неё нет: колонки называются то «Тема урока», то
«Содержание», то «Раздел/тема», заголовок стоит то первой строкой, то
пятой, а сверху бывает шапка с грифами и подписями.

Поэтому разбор устроен в два шага. Сначала таблица читается как есть —
строки и ячейки, без всякого смысла. Потом по названиям колонок угадывается,
что где лежит. Угаданное показывается человеку и правится руками: угадать
чужую таблицу с первого раза нельзя, а вот показать, что понял, и дать
поправить — можно всегда.

Когда придёт образец, править нужно будет ровно одно место — SYNONYMS
ниже. Всё остальное от вида таблицы не зависит.
"""
from __future__ import annotations

import csv
import datetime as dt
import io
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.exceptions import ValidationError

# Поля строки КТП. Порядок — тот, в котором их показывают человеку.
FIELDS = [
    ("number", "№ по плану"),
    ("planned_date", "Дата"),
    ("topic", "Тема"),
    ("hours", "Часов"),
    ("kind", "Тип занятия"),
    ("homework", "Домашнее задание"),
    ("notes", "Примечание"),
]
FIELD_NAMES = [name for name, _ in FIELDS]

# Как называют эти колонки в живых таблицах. Сравнение идёт по началу
# строки и по вхождению, в нижнем регистре, без пунктуации: «Тема урока»,
# «тема  занятия», «ТЕМА» — одно и то же.
SYNONYMS: dict[str, tuple[str, ...]] = {
    "number": ("n п/п", "n урока", "n занятия", "номер", "n"),
    "planned_date": ("дата план", "план дата", "дата проведения", "дата по плану",
                     "дата", "план", "сроки"),
    "topic": ("тема урока", "тема занятия", "тема", "содержание", "раздел тема", "наименование"),
    "hours": ("количество часов", "кол во часов", "часов", "часы", "ч"),
    "kind": ("тип урока", "тип занятия", "форма занятия", "форма", "вид занятия", "тип"),
    "homework": ("домашнее задание", "дом задание", "д з", "дз", "задание на дом"),
    "notes": ("примечание", "примечания", "оборудование", "средства обучения",
              "планируемые результаты", "ууд", "контроль"),
}

# Слова, при которых колонка точно не про это поле. «Дата факт» стоит
# рядом с «Датой план» и по одному слову «дата» неотличима от неё — а
# перепутать их значит разложить план по тому, что ещё не случилось.
DISQUALIFIERS: dict[str, tuple[str, ...]] = {
    "planned_date": ("факт",),
}

SUPPORTED_SUFFIXES = {".xlsx", ".xlsm", ".csv"}
MAX_PREVIEW_ROWS = 400
# Заголовок ищем только в начале файла: ниже начинаются сами занятия, и
# строка «Тема» там — это тема, а не название колонки.
HEADER_SEARCH_DEPTH = 12


@dataclass
class ParsedRow:
    """Разобранная строка КТП. Пустые поля — это пустые поля, не ошибка."""

    number: str = ""
    planned_date: dt.date | None = None
    topic: str = ""
    hours: Decimal = Decimal("1.00")
    kind: str = ""
    homework: str = ""
    notes: str = ""
    # Название раздела: «Причастие 21ч +2К + 2Р.р». Занятием не является —
    # ни часов, ни даты у него нет, только заголовок над группой строк.
    is_section: bool = False

    @property
    def is_empty(self) -> bool:
        return not self.topic.strip()


@dataclass
class ParsedPlan:
    """Что вышло из файла: сырые строки, разметка колонок и разобранное."""

    table: list[list[str]] = field(default_factory=list)
    header_row: int = 0
    headers: list[str] = field(default_factory=list)
    column_map: dict[str, int] = field(default_factory=dict)
    rows: list[ParsedRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


# ─── Чтение файла ───────────────────────────────────────────────────────────

def read_table(uploaded, *, filename: str = "") -> list[list[str]]:
    """
    Файл → строки ячеек. Ничего не интерпретируем: что написано, то и берём.

    Дата из таблицы приходит датой, а не строкой, — её оставляем в
    исходном виде, чтобы потом не разбирать заново то, что уже разобрано.
    """
    name = (filename or getattr(uploaded, "name", "") or "").lower()
    suffix = Path(name).suffix
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValidationError(
            "Такой файл прочитать нельзя. Подойдёт таблица Excel (.xlsx) "
            "или .csv — обычно КТП присылают именно так."
        )

    uploaded.seek(0)
    if suffix == ".csv":
        return _read_csv(uploaded.read())
    return _read_xlsx(uploaded)


def _read_csv(payload: bytes) -> list[list[str]]:
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            text = payload.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValidationError("Не удалось прочитать файл: неизвестная кодировка.")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    return [list(row) for row in csv.reader(io.StringIO(text), dialect)]


def _read_xlsx(uploaded) -> list[list[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(uploaded, data_only=True, read_only=True)
    sheet = workbook.active
    table: list[list[str]] = []
    for raw in sheet.iter_rows(values_only=True):
        table.append(["" if cell is None else cell for cell in raw])
    workbook.close()
    return table


# ─── Разметка колонок ───────────────────────────────────────────────────────

def _normalize(value) -> str:
    """
    Заголовок к сравнимому виду.

    «№» превращаем в «n», а не выбрасываем вместе с пунктуацией: иначе
    синоним «№» станет пустой строкой, а пустая строка входит в любую —
    и колонка с номером находилась бы где угодно, хоть в теме.
    """
    text = str(value or "").strip().lower().replace("ё", "е").replace("№", " n ")
    return re.sub(r"[^a-zа-я0-9 ]+", " ", text).strip()


def _score(header: str, field_name: str) -> int:
    """Насколько заголовок похож на колонку. Точное совпадение весомее вхождения."""
    header = _normalize(header)
    if not header:
        return 0
    if any(word in header for word in DISQUALIFIERS.get(field_name, ())):
        return 0
    best = 0
    for synonym in SYNONYMS[field_name]:
        synonym = _normalize(synonym)
        if not synonym:
            continue
        if header == synonym:
            best = max(best, 100)
        elif header.startswith(synonym) or synonym in header:
            # Длинный синоним, найденный внутри, надёжнее короткого.
            best = max(best, 40 + len(synonym))
    return best


def guess_header(table: list[list[str]]) -> tuple[int, dict[str, int]]:
    """
    Найти строку заголовка и понять, какая колонка что значит.

    Заголовком считаем ту строку в начале файла, которая опознаётся лучше
    прочих: у настоящего заголовка совпадают сразу несколько колонок, у
    случайной строки — в лучшем случае одна.
    """
    best_row, best_map, best_total = 0, {}, 0
    for index, row in enumerate(table[:HEADER_SEARCH_DEPTH]):
        mapping, total = _map_row(row)
        # Тема — колонка, без которой плана нет: строка без неё заголовком
        # быть не может, как бы хорошо ни совпало остальное.
        if "topic" not in mapping:
            continue
        if total > best_total:
            best_row, best_map, best_total = index, mapping, total
    return best_row, best_map


def _map_row(row: list[str]) -> tuple[dict[str, int], int]:
    """
    Разобрать строку заголовка.

    Одна колонка достаётся одному полю: «Тема» — это тема, и заодно
    номером быть не может. Побеждает та пара, у которой совпадение
    увереннее, — так колонка не растаскивается по двум смыслам.
    """
    scored = [
        (_score(header, name), name, column)
        for column, header in enumerate(row)
        for name in FIELD_NAMES
        if _score(header, name)
    ]
    scored.sort(key=lambda item: -item[0])

    mapping: dict[str, int] = {}
    used: set[int] = set()
    total = 0
    for score, name, column in scored:
        if name in mapping or column in used:
            continue
        mapping[name] = column
        used.add(column)
        total += score
    return mapping, total


# ─── Разбор значений ────────────────────────────────────────────────────────

def parse_date(value) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    text = text.split()[0]
    for pattern in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m"):
        try:
            parsed = dt.datetime.strptime(text, pattern).date()
        except ValueError:
            continue
        # «12.09» без года встречается сплошь и рядом; год подставит
        # привязка к расписанию, а пока такая дата бесполезна.
        return None if pattern == "%d.%m" else parsed
    return None


def parse_hours(value) -> Decimal:
    text = str(value or "").strip().replace(",", ".")
    if not text:
        return Decimal("1.00")
    match = re.search(r"\d+(\.\d+)?", text)
    if not match:
        return Decimal("1.00")
    try:
        hours = Decimal(match.group(0))
    except InvalidOperation:
        return Decimal("1.00")
    return hours if Decimal("0") < hours <= Decimal("99") else Decimal("1.00")


def _cell(row: list[str], column: int | None) -> str:
    if column is None or column >= len(row):
        return ""
    return str(row[column] or "").strip()


def parse(
    table: list[list[str]], *, header_row: int | None = None, column_map: dict | None = None
) -> ParsedPlan:
    """
    Разобрать таблицу по разметке колонок. Разметку можно задать снаружи —
    ровно для того, чтобы человек мог поправить угаданное.
    """
    plan = ParsedPlan(table=table)
    if not table:
        plan.warnings.append("Файл пустой.")
        return plan

    if column_map:
        plan.column_map = {k: int(v) for k, v in column_map.items() if v not in (None, "")}
        plan.header_row = header_row or 0
    else:
        plan.header_row, plan.column_map = guess_header(table)
        if header_row is not None:
            plan.header_row = header_row

    plan.headers = [str(cell or "") for cell in table[min(plan.header_row, len(table) - 1)]]
    if "topic" not in plan.column_map:
        plan.warnings.append(
            "Не нашли колонку с темой. Укажите её вручную — без темы план не собрать."
        )
        return plan

    # Раздел от занятия отличаем по часам: у занятия они есть, у заголовка
    # раздела нет. Работает это, только когда колонка с часами вообще есть,
    # — иначе разделами оказался бы весь план.
    hours_column = plan.column_map.get("hours")

    for raw in table[plan.header_row + 1 :]:
        hours_cell = _cell(raw, hours_column)
        row = ParsedRow(
            number=_cell(raw, plan.column_map.get("number"))[:20],
            planned_date=parse_date(
                raw[plan.column_map["planned_date"]]
                if "planned_date" in plan.column_map and plan.column_map["planned_date"] < len(raw)
                else ""
            ),
            topic=_cell(raw, plan.column_map.get("topic"))[:250],
            hours=parse_hours(hours_cell),
            kind=_cell(raw, plan.column_map.get("kind"))[:80],
            homework=_cell(raw, plan.column_map.get("homework")),
            notes=_cell(raw, plan.column_map.get("notes")),
        )
        if row.is_empty:
            # Пустая строка внутри таблицы — обычно разделитель раздела.
            continue
        row.is_section = hours_column is not None and not hours_cell
        plan.rows.append(row)
        if len(plan.rows) >= MAX_PREVIEW_ROWS:
            plan.warnings.append(
                f"Показаны первые {MAX_PREVIEW_ROWS} строк — остальные не разбирались."
            )
            break

    if not plan.rows:
        plan.warnings.append("Строк с темами не нашлось. Проверьте, ту ли колонку выбрали темой.")
    elif all(row.is_section for row in plan.rows):
        plan.warnings.append(
            "Ни у одной строки нет часов — похоже, колонка с часами выбрана неверно."
        )
    return plan


# ─── Сохранение и привязка к расписанию ─────────────────────────────────────

def save_entries(plan_object, parsed: ParsedPlan) -> int:
    """
    Записать разобранные строки, заменив прежние.

    Заменяем целиком, а не дописываем: разбор повторяют, когда предыдущий
    оказался неверным, и дописанное к неверному хуже, чем неверное.
    """
    from apps.journal.models import ThematicPlanEntry

    ThematicPlanEntry.objects.filter(plan=plan_object).delete()
    ThematicPlanEntry.objects.bulk_create(
        [
            ThematicPlanEntry(
                organization=plan_object.organization, plan=plan_object, position=index,
                number=row.number, planned_date=row.planned_date, topic=row.topic,
                hours=row.hours, kind=row.kind, homework=row.homework, notes=row.notes,
                is_section=row.is_section,
            )
            for index, row in enumerate(parsed.rows, start=1)
        ]
    )
    plan_object.header_row = parsed.header_row
    plan_object.column_map = parsed.column_map
    plan_object.save(update_fields=["header_row", "column_map", "updated_at"])
    return len(parsed.rows)


def attach_to_lessons(plan_object, *, overwrite: bool = False) -> dict[str, int]:
    """
    Разложить темы плана по занятиям расписания.

    Сначала по дате — если в плане она есть и в этот день занятие одно.
    Остальное раскладывается по порядку: n-я строка плана на n-е занятие
    предмета. Занятия с уже вписанной темой не трогаем, пока не попросят
    прямо: тема, которую педагог дописал руками, точнее плановой.
    """
    from apps.journal.models import Lesson

    lessons = list(
        Lesson.objects.filter(
            subject=plan_object.subject, module__academic_year=plan_object.academic_year
        )
        .filter(**({"group": plan_object.group} if plan_object.group_id else {}))
        .order_by("starts_at")
    )
    # Заголовки разделов занятиями не являются — раскладывать их некуда.
    entries = list(plan_object.entries.filter(is_section=False).order_by("position"))

    by_date: dict[dt.date, list] = {}
    for lesson in lessons:
        by_date.setdefault(lesson.local_date, []).append(lesson)

    used: set = set()
    result = {"matched": 0, "filled": 0, "skipped": 0}

    def take(entry, lesson):
        entry.lesson = lesson
        used.add(lesson.pk)
        result["matched"] += 1
        if lesson.topic and not overwrite:
            result["skipped"] += 1
            return
        if lesson.topic != entry.topic:
            lesson.topic = entry.topic
            lesson.save(update_fields=["topic", "updated_at"])
            result["filled"] += 1

    leftovers = []
    for entry in entries:
        same_day = [
            lesson for lesson in by_date.get(entry.planned_date, []) if lesson.pk not in used
        ] if entry.planned_date else []
        if len(same_day) == 1:
            take(entry, same_day[0])
        else:
            leftovers.append(entry)

    free = [lesson for lesson in lessons if lesson.pk not in used]
    for entry, lesson in zip(leftovers, free):
        take(entry, lesson)

    from apps.journal.models import ThematicPlanEntry

    ThematicPlanEntry.objects.bulk_update(entries, ["lesson"])
    result["unmatched"] = sum(1 for entry in entries if entry.lesson_id is None)
    return result


# ─── Связь темы занятия с планом ────────────────────────────────────────────

def entry_for(lesson):
    """
    Строка КТП, привязанная к занятию, — если план разложен по расписанию.

    Заголовки разделов сюда не попадают: они не занятия.
    """
    from apps.journal.models import ThematicPlanEntry

    return (
        ThematicPlanEntry.objects.filter(lesson=lesson, is_section=False)
        .select_related("plan", "plan__subject")
        .first()
    )


def sync_topic_from_lesson(lesson) -> bool:
    """
    Тема, исправленная в занятии, уходит обратно в КТП.

    План — не высеченный в камне документ: педагог правит формулировку по
    ходу года, и держать в плане одно, а в журнале другое значит завести
    два разных плана. Правка идёт в одну сторону — от занятия к плану, —
    потому что занятие всегда конкретнее: оно уже состоялось.
    """
    from apps.journal.models import ThematicPlanEntry

    topic = (lesson.topic or "").strip()[:250]
    if not topic:
        # Стёртая тема занятия не стирает плановую: у плана она была
        # осмысленной, а пустое поле — это чаще всего «ещё не заполнил».
        return False
    return bool(
        ThematicPlanEntry.objects.filter(lesson=lesson, is_section=False)
        .exclude(topic=topic)
        .update(topic=topic)
    )
