"""
Разбор расписания, которое ведётся в таблице.

Заказчик ведёт расписание сеткой: слева время, сверху дни, в клетках —
названия занятий. Просить переводить это в CSV каждый раз — способ
получить ошибку переноса, поэтому таблица читается как есть.

Здесь только разбор: ни одного обращения к базе. Так формат можно
проверять тестами на голых строках, не поднимая организацию.
"""
from __future__ import annotations

import datetime as dt
import io
import re
import zipfile
from dataclasses import dataclass

WEEKDAY_LABELS = {
    "пн": 0, "понедельник": 0,
    "вт": 1, "вторник": 1,
    "ср": 2, "среда": 2,
    "чт": 3, "четверг": 3,
    "пт": 4, "пятница": 4,
    "сб": 5, "суббота": 5,
    "вс": 6, "воскресенье": 6,
}

# Как в таблице называются вещи, которые в журнале называются иначе.
# Список короткий намеренно: каждое расхождение лучше однажды исправить
# в таблице, чем прятать в коде.
ALIASES = {
    "вист": "Вероятность и статистика",
    "аглебра": "Алгебра",  # опечатка в исходном файле
    "профориентация/самоподготовка": "Профориентация",
}

# Клетки, которые не являются занятием вообще.
IGNORED = {"", "-", "—"}

DAY_HEADER_RE = re.compile(
    r"^\s*(?P<label>[А-Яа-яЁё]{2,12})\.?\s*(?P<day>\d{1,2})[.,/](?P<month>\d{1,2})"
)
# Заголовок может быть и наоборот: «02.09 Ср».
DATE_FIRST_RE = re.compile(
    r"^\s*(?P<day>\d{1,2})[.,/](?P<month>\d{1,2})\.?\s*(?P<label>[А-Яа-яЁё]{2,12})"
)
TIME_RANGE_RE = re.compile(
    r"^\s*(?P<h1>\d{1,2})[.:](?P<m1>\d{2})\s*[-–—]\s*(?P<h2>\d{1,2})[.:](?P<m2>\d{2})"
)
# «1. 9.30-10.10» — номер урока перед временем.
LESSON_NUMBER_RE = re.compile(r"^\s*\d{1,2}\s*[.)]\s*")


@dataclass(frozen=True)
class ParsedLesson:
    date: dt.date
    start: dt.time
    duration_minutes: int
    title: str
    column_label: str


@dataclass(frozen=True)
class ParseResult:
    lessons: list[ParsedLesson]
    warnings: list[str]


def load_workbook_rows(path) -> dict[str, list[list]]:
    """
    Прочитать книгу в простые списки значений.

    Яндекс.Документы записывают в стилях `style="solid"` — такого значения
    в формате нет, и openpyxl отказывается открывать файл целиком. Чиним
    копию в памяти: сам файл заказчика при этом не трогаем.
    """
    import openpyxl

    with open(path, "rb") as handle:
        raw = handle.read()

    repaired = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(raw)) as zin:
        with zipfile.ZipFile(repaired, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/styles.xml":
                    data = data.replace(b'style="solid"', b'style="thin"')
                zout.writestr(item, data)
    repaired.seek(0)

    # Не read_only: в этих книгах не заполнен блок с размерами листа,
    # и в потоковом режиме openpyxl отдаёт одну строку вместо всей таблицы.
    workbook = openpyxl.load_workbook(repaired, data_only=True)
    return {
        sheet.title: [list(row) for row in sheet.iter_rows(values_only=True)]
        for sheet in workbook.worksheets
    }


def parse_day_header(value) -> tuple[int, int, int] | None:
    """«ПН 30.08» или «02.09 Ср» → (день недели, число, месяц)."""
    if value is None:
        return None
    text = str(value).strip()
    for pattern in (DAY_HEADER_RE, DATE_FIRST_RE):
        match = pattern.match(text)
        if not match:
            continue
        label = match.group("label").lower().rstrip(".")
        if label not in WEEKDAY_LABELS:
            continue
        return WEEKDAY_LABELS[label], int(match.group("day")), int(match.group("month"))
    return None


def parse_time_range(value) -> tuple[dt.time, int] | None:
    """«9.30-10.10» → начало 9:30 и 40 минут. Номер урока перед временем допустим."""
    if value is None:
        return None
    text = str(value).strip()
    match = TIME_RANGE_RE.match(text)
    if not match:
        # Номер урока снимаем только второй попыткой: «9.30-10.10» само по себе
        # начинается с того же, с чего начинается «9. …», и снятое вслепую
        # превратило бы время в мусор.
        match = TIME_RANGE_RE.match(LESSON_NUMBER_RE.sub("", text))
    if not match:
        return None
    start = dt.time(int(match.group("h1")), int(match.group("m1")))
    end = dt.time(int(match.group("h2")), int(match.group("m2")))
    minutes = (end.hour * 60 + end.minute) - (start.hour * 60 + start.minute)
    return start, minutes if minutes > 0 else 40


def normalize_title(value) -> str:
    """Схлопнуть лишние пробелы. Название из таблицы остаётся как есть."""
    return " ".join(str(value or "").split())


def title_candidates(value) -> list[str]:
    """
    Как искать это название в журнале: сначала как написано, потом синоним.

    Синоним второй, а не первый, намеренно. «Математика» у этого заказчика
    разложена на алгебру с геометрией, но у другой организации предмет
    с таким названием может существовать — и тогда подменять его нельзя.
    """
    title = normalize_title(value)
    alias = ALIASES.get(title.lower())
    return [title] if alias is None else [title, alias]


def _resolve_year(day: int, month: int, weekday: int, within: tuple[dt.date, dt.date]) -> tuple[dt.date, str | None]:
    """
    Год в сетке не пишут — берём его из учебного года.

    Если день недели в подписи не сходится с датой, дату не подменяем:
    в таблице бывает и опечатка в подписи, и опечатка в числе, и угадывать,
    какая именно, — плохая идея. Возвращаем предупреждение.
    """
    starts_on, ends_on = within
    candidates = []
    for year in {starts_on.year, ends_on.year}:
        try:
            candidates.append(dt.date(year, month, day))
        except ValueError:
            continue
    if not candidates:
        return None, f"невозможная дата {day:02d}.{month:02d}"

    inside = [d for d in candidates if starts_on <= d <= ends_on]
    date = min(inside or candidates, key=lambda d: abs((d - starts_on).days))
    if date.weekday() != weekday:
        names = ["понедельник", "вторник", "среду", "четверг", "пятницу", "субботу", "воскресенье"]
        return date, (
            f"{date:%d.%m.%Y} — это {names[date.weekday()]}, "
            f"а в таблице подписано как {names[weekday]}"
        )
    return date, None


def parse_grid(rows: list[list], *, within: tuple[dt.date, dt.date]) -> ParseResult:
    """
    Разобрать сетку «время слева, дни сверху».

    Блоков в листе может быть несколько (неделя за неделей) — каждая строка
    с датами в шапке начинает новый блок.
    """
    lessons: list[ParsedLesson] = []
    warnings: list[str] = []
    columns: dict[int, tuple[dt.date, str]] = {}
    seen_dates: dict[dt.date, int] = {}

    for row_number, row in enumerate(rows, start=1):
        header = {}
        for index, cell in enumerate(row):
            parsed = parse_day_header(cell)
            if parsed is not None:
                weekday, day, month = parsed
                date, warning = _resolve_year(day, month, weekday, within)
                if warning:
                    warnings.append(f"строка {row_number}: {warning}")
                if date is not None:
                    header[index] = (date, str(cell).strip())

        if len(header) >= 2:
            for date, label in header.values():
                first = seen_dates.get(date)
                if first is not None:
                    # Верный признак копипасты в шапке: одна и та же дата
                    # в двух блоках. Молча слить их — потерять целый день,
                    # поэтому говорим прямо.
                    warnings.append(
                        f"строка {row_number}: дата «{label}» уже была в строке {first}. "
                        "Похоже, шапку скопировали и не поправили — день недели "
                        "из второго блока никуда не попадёт"
                    )
                else:
                    seen_dates[date] = row_number
            columns = header
            continue

        if not columns:
            continue

        moment = None
        for cell in row[: min(columns)]:
            moment = parse_time_range(cell)
            if moment is not None:
                break
        if moment is None:
            continue
        start, duration = moment

        for index, (date, label) in columns.items():
            if index >= len(row):
                continue
            title = normalize_title(row[index])
            if title.lower() in IGNORED:
                continue
            lessons.append(
                ParsedLesson(
                    date=date, start=start, duration_minutes=duration,
                    title=title, column_label=label,
                )
            )

    return ParseResult(lessons=lessons, warnings=warnings)


def pick_sheet(sheets: dict[str, list[list]], preferred: str | None = None) -> str:
    """Лист с расписанием: по имени, иначе тот, где больше всего дат в шапке."""
    if preferred:
        if preferred not in sheets:
            raise KeyError(preferred)
        return preferred
    named = [title for title in sheets if "расписан" in title.lower()]
    if named:
        return named[0]

    def day_headers(rows) -> int:
        return sum(
            1 for row in rows for cell in row if parse_day_header(cell) is not None
        )

    return max(sheets, key=lambda title: day_headers(sheets[title]))
