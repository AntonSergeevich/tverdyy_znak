"""
Экспорт в xlsx (ТЗ 7).

Важное правило: скрытые личные цели ученика не попадают ни в одну выгрузку.
Все функции здесь берут цели только через Goal.objects.visible_to_others().
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from apps.journal.models import (
    Goal,
    GoalStatus,
    ModuleResult,
    Student,
    StudentStatus,
)

HEADER_FONT = Font(bold=True)


def _finish(workbook: Workbook) -> bytes:
    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            width = max((len(str(cell.value or "")) for cell in column_cells), default=10)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
                48, max(12, width + 2)
            )
        sheet.freeze_panes = "A2"
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _header(sheet, titles: list[str]) -> None:
    sheet.append(titles)
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def module_performance_xlsx(module) -> bytes:
    """Успеваемость по модулю: ученик × предмет, сумма баллов и уровень."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Модуль {module.number}"[:31]
    _header(sheet, ["Ученик", "Класс", "Предмет", "Баллы", "Из", "Уровень", "Зачёт"])

    results = (
        ModuleResult.objects.filter(module=module)
        .select_related("student", "subject")
        .order_by("student__last_name", "subject__position")
    )
    for result in results:
        sheet.append(
            [
                result.student.full_name,
                result.student.grade_level,
                result.subject.name,
                float(result.total_points),
                float(result.planned_points),
                result.get_level_display(),
                "да" if result.is_passed else "нет",
            ]
        )
    return _finish(workbook)


def students_xlsx() -> bytes:
    """
    Список учеников.

    Дата рождения и документы намеренно не выгружаются: это чувствительные
    данные несовершеннолетних, и в файле, который уходит из системы,
    им делать нечего.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ученики"
    _header(sheet, ["Фамилия", "Имя", "Отчество", "Класс", "Статус", "Зачислен", "Родители"])

    students = (
        Student.objects.all()
        .prefetch_related("parent_links__parent")
        .order_by("last_name", "first_name")
    )
    for student in students:
        parents = ", ".join(link.parent.full_name for link in student.parent_links.all())
        sheet.append(
            [
                student.last_name,
                student.first_name,
                student.middle_name,
                student.grade_level,
                student.get_status_display(),
                student.enrolled_on.isoformat() if student.enrolled_on else "",
                parents,
            ]
        )
    return _finish(workbook)


def payroll_xlsx(rows: list[dict], start: date, end: date) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ФОТ"
    # Часы академические: в бухгалтерию должна уходить та же цифра, что
    # педагог видит у себя, иначе разбираться будут не с таблицей, а друг
    # с другом.
    _header(sheet, ["Педагог", "Занятий", "Академических часов", "Ставка", "К выплате"])
    total = Decimal("0.00")
    for row in rows:
        sheet.append(
            [
                row["teacher"].user.full_name,
                row.get("lessons", 0),
                row["hours"],
                float(row["rate"]),
                float(row["amount"]),
            ]
        )
        total += row["amount"]
    sheet.append(["Итого", "", "", "", float(total)])
    sheet.append([f"Период: {start.isoformat()} — {end.isoformat()}"])
    return _finish(workbook)


def goals_xlsx(student: Student) -> bytes:
    """
    Цели ученика для наставника и родителя.

    Скрытые личные цели сюда не попадают — это требование ТЗ 5.2
    и предмет отдельного теста.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Цели"
    _header(sheet, ["Тип", "Предмет", "Цель", "Срок", "Статус"])

    goals = (
        Goal.objects.visible_to_others()
        .filter(student=student)
        .select_related("subject")
        .order_by("target_date")
    )
    for goal in goals:
        sheet.append(
            [
                goal.get_kind_display(),
                goal.subject.name if goal.subject_id else "",
                goal.title,
                goal.target_date.isoformat() if goal.target_date else "",
                goal.get_status_display(),
            ]
        )
    return _finish(workbook)
